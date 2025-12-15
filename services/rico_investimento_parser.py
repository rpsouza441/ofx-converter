#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Parser para arquivos XLSX de investimentos da Rico
Formato: Movimentação | Liquidação | Lançamento | Qtd | Valor (R$) | Saldo (R$)
"""

import logging
from datetime import datetime
from typing import Dict, List, Any, Optional
import openpyxl
from services.text_normalizer import TextNormalizer
from services.categorizer import TransactionCategorizer

logger = logging.getLogger(__name__)


class RicoInvestimentoParser:
    """Parser para extratos XLSX de investimentos da Rico"""
    
    def __init__(self, categorizer: TransactionCategorizer):
        self.text_normalizer = TextNormalizer()
        self.categorizer = categorizer
    
    def parse(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Faz parse de arquivo XLSX de investimentos da Rico
        
        Args:
            file_path: Caminho do arquivo XLSX
            
        Returns:
            Lista de transações parseadas
        """
        transactions = []
        
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Encontrar linha do header (procurar por "Movimentação")
            header_row = None
            for row_num in range(1, min(20, ws.max_row + 1)):
                cell_value = ws.cell(row=row_num, column=1).value
                if cell_value and 'Movimentação' in str(cell_value):
                    header_row = row_num
                    break
            
            if not header_row:
                logger.error(f"Header 'Movimentação' não encontrado no XLSX: {file_path}")
                return []
            
            logger.info(f"Header encontrado na linha {header_row}")
            
            # Processar linhas de dados (após header)
            for row_num in range(header_row + 1, ws.max_row + 1):
                try:
                    # Ler células da linha
                    liquidacao = ws.cell(row=row_num, column=2).value  # Data liquidação
                    lancamento = ws.cell(row=row_num, column=3).value  # Descrição
                    # Coluna 4 = quantidade (informativo, ignorar)
                    valor = ws.cell(row=row_num, column=5).value       # Valor R$
                    # Coluna 6 = saldo (ignorar)
                    
                    # Pular linhas vazias
                    if not liquidacao or not lancamento:
                        continue
                    
                    transaction = self._parse_transaction(
                        liquidacao, lancamento, valor
                    )
                    
                    if transaction:
                        transactions.append(transaction)
                        
                except Exception as e:
                    logger.warning(f"Erro ao parsear linha {row_num}: {e}")
                    continue
            
            logger.info(f"Rico Investimento: {len(transactions)} transações parseadas de {file_path}")
            return transactions
            
        except Exception as e:
            logger.error(f"Erro ao processar XLSX Rico Investimento {file_path}: {e}")
            raise
    
    def _parse_transaction(self, liquidacao, lancamento: str, 
                          valor) -> Optional[Dict[str, Any]]:
        """
        Parseia uma transação de investimentos da Rico
        
        Args:
            liquidacao: Data de liquidação (datetime ou string) - quando $ movimenta
            lancamento: Descrição da transação (ex: RENDIMENTOS DE CLIENTES BTLG11)
            valor: Valor (string "R$ X,XX" ou número)
        """
        try:
            # Parse data de liquidação
            date_str = self._parse_date(liquidacao)
            
            # Parse descrição
            description = self.text_normalizer.normalize_utf8(str(lancamento).strip())
            
            # Parse valor
            amount = self._parse_amount(valor)
            
            # Categorizar
            category_info = self.categorizer.categorize_smart(description, amount)
            
            return {
                'date': date_str,
                'description': description,
                'amount': amount,
                'type': category_info['type'],
                'category': category_info['category'],
                'subcategory': category_info.get('subcategory', ''),
                'qif_category': self._get_qif_category(category_info)
            }
            
        except Exception as e:
            logger.error(f"Erro ao parsear transação: {e}")
            return None
    
    def _parse_date(self, date_value) -> str:
        """
        Parse data que pode vir como datetime object ou string DD/MM/YYYY
        
        Returns:
            String no formato 'YYYY-MM-DD 00:00:00'
        """
        try:
            if isinstance(date_value, datetime):
                return date_value.strftime('%Y-%m-%d 00:00:00')
            elif isinstance(date_value, str):
                # Parse DD/MM/YYYY
                dt = datetime.strptime(date_value, '%d/%m/%Y')
                return dt.strftime('%Y-%m-%d 00:00:00')
            else:
                logger.error(f"Formato de data desconhecido: {type(date_value)}")
                return datetime.now().strftime('%Y-%m-%d 00:00:00')
        except Exception as e:
            logger.error(f"Erro ao parsear data '{date_value}': {e}")
            return datetime.now().strftime('%Y-%m-%d 00:00:00')
    
    def _parse_amount(self, value) -> float:
        """
        Parse valor que pode vir como:
        - String: "R$ 1.234,56" ou "-R$ 19,84"
        - Float direto
        
        Returns:
            Float com valor (positivo ou negativo)
        """
        try:
            if isinstance(value, (int, float)):
                return float(value)
            
            if isinstance(value, str):
                # Remover "R$" e espaços
                clean = value.replace('R$', '').strip()
                
                # Verificar sinal negativo
                is_negative = clean.startswith('-')
                clean = clean.replace('-', '').strip()
                
                # Remover pontos (milhares) e trocar vírgula por ponto (decimais)
                clean = clean.replace('.', '').replace(',', '.')
                
                amount = float(clean)
                return -amount if is_negative else amount
            
            return 0.0
            
        except Exception as e:
            logger.error(f"Erro ao parsear valor '{value}': {e}")
            return 0.0
    
    def _get_qif_category(self, category_info: dict) -> str:
        """
        Gera categoria QIF com base no tipo
        
        Args:
            category_info: Dict com type, category, subcategory
            
        Returns:
            String de categoria QIF
        """
        if category_info['type'] == 'transfer':
            return f"[{category_info['category']}]"
        else:
            return category_info['category']
